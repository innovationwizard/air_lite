"""
Convert account.move.line_2026.xlsx to CSV.
One-time utility — run before ingest.py.
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print('openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)


def convert(xlsx_path: str, csv_path: str) -> None:
    import csv

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([cell if cell is not None else '' for cell in row])

    wb.close()
    print(f'Converted {xlsx_path} -> {csv_path}')


if __name__ == '__main__':
    real_data = Path(__file__).resolve().parent.parent / 'real_data'
    xlsx_file = real_data / 'account.move.line_2026.xlsx'
    csv_file = real_data / 'account.move.line_2026.csv'

    if not xlsx_file.exists():
        print(f'File not found: {xlsx_file}')
        sys.exit(1)

    convert(str(xlsx_file), str(csv_file))
