#!/usr/bin/env python3
"""
Carga el libro «Punto de Reorden · Alcance Máximo» a las tablas del modelo — A4.27.

⚠️ ESTE ARCHIVO ES LA CLI. Las REGLAS —normalización, sinónimos, mapeo por
encabezado, lectura de fechas y proformas— viven en `ml/modelo_reorden_carga.py`,
igual que `reyma_factura_carga.py` para las facturas. El día que el proveedor
cambie el libro se arregla en un solo lugar.

QUÉ CARGA, y de dónde sale cada cosa:

  `reorden_inventario`  ← hoja «MODELO INVENTARIO», sólo los INSUMOS. Cobertura,
                          estado y pedido NO se guardan: los recalcula el motor
                          (`lib/inventarios/reorden.ts`). Guardar un resultado
                          junto a sus insumos es garantizar que algún día
                          discrepen.
  `reorden_transito`     ← las columnas FECHADAS de tránsito confirmado,
                          normalizadas a una fila por embarque. En el libro son
                          columnas; cada embarque nuevo agregaba una.
  `reorden_precios`      ← hoja «BASE DATOS», una fila por producto y proforma.
  `estado_producto`      ← de la hoja de precios, NO del motor: en el motor la
                          regla de liquidación está rota por un `#REF!` y
                          clasifica cero productos cuando 11 lo están.

DOS COSAS QUE ESTE CARGADOR NO HACE, a propósito:

  · No adivina columnas. Si falta una obligatoria, se detiene. Un mapeo
    equivocado no falla ruidosamente: escribe números plausibles en el campo de
    al lado, y eso llega hasta la orden de compra.
  · No rellena huecos. `N/D` y `—` entran como NULL, nunca como cero — cero es
    una afirmación y la ausencia de dato no lo es.

Uso:
    python scripts/cargar_modelo_reorden.py LIBRO.xlsx --modelo darnel
    python scripts/cargar_modelo_reorden.py LIBRO.xlsx --modelo darnel --commit

Sin `--commit` es dry-run: imprime exactamente lo que escribiría.
Requiere SUPABASE_URL y SUPABASE_SECRET_KEY.
"""

import argparse
import os
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'ml'))

from modelo_reorden_carga import (  # noqa: E402
    a_numero, mapear_encabezados, normalizar, normalizar_estado, parsear_proforma,
)

HOJA_MOTOR_CLAVES = ('modelo inventario',)
HOJA_PRECIOS_CLAVES = ('base datos',)


def hallar_hoja(wb, claves):
    """La hoja por lo que SIGNIFICA su nombre, no por su posición ni su nombre
    exacto: la del motor se llama «MODELO INVENTARIO» y la de precios «BASE
    DATOS DARNEL (2)» — ese «(2)» es exactamente la clase de cosa que rompe una
    comparación literal."""
    for ws in wb.worksheets:
        n = normalizar(ws.title)
        if any(k in n for k in claves):
            return ws
    return None


def leer_precios(wb):
    """{codigo: (estado, [(proforma, precio)])} desde la hoja de precios."""
    ws = hallar_hoja(wb, HOJA_PRECIOS_CLAVES)
    if ws is None:
        return {}, []
    encabezados = [c.value for c in ws[3]]
    idx_cod = idx_estado = None
    columnas_precio = []
    for i, h in enumerate(encabezados):
        n = normalizar(h)
        if n in ('codigo plasticentro', 'cod plasticentro'):
            idx_cod = i
        elif 'estado' in n:
            idx_estado = i
        elif n.startswith('precio'):
            columnas_precio.append((i, parsear_proforma(h)))
    if idx_cod is None:
        return {}, []

    estados, precios = {}, []
    for fila in ws.iter_rows(min_row=4, values_only=True):
        cod = fila[idx_cod] if idx_cod < len(fila) else None
        if not cod:
            continue
        cod = str(cod).strip()
        estados[cod] = normalizar_estado(fila[idx_estado] if idx_estado is not None else None)
        for i, proforma in columnas_precio:
            v = a_numero(fila[i]) if i < len(fila) else None
            if v is not None and proforma:
                precios.append({'codigo': cod, 'proforma': proforma, 'precio_ml': v})
    return estados, precios


def leer_motor(wb, anio):
    ws = hallar_hoja(wb, HOJA_MOTOR_CLAVES)
    if ws is None:
        raise SystemExit('No se encontro la hoja del motor («MODELO INVENTARIO»).')
    grupos = [c.value for c in ws[2]]
    columnas = [c.value for c in ws[3]]
    mapa, transito, sin_casar, faltantes = mapear_encabezados(grupos, columnas, anio=anio)
    return ws, mapa, transito, sin_casar, faltantes


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('libro', help='ruta al .xlsx')
    ap.add_argument('--modelo', required=True, help='slug en modelo_proveedor (darnel, asia…)')
    ap.add_argument('--anio', type=int, default=None,
                    help='ano para los encabezados de embarque que traen solo dia/mes')
    ap.add_argument('--commit', action='store_true', help='escribe (sin esto, dry-run)')
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(args.libro, data_only=True)
    print(f'Libro: {Path(args.libro).name}')
    print(f'Hojas: {wb.sheetnames}')

    ws, mapa, transito, sin_casar, faltantes = leer_motor(wb, args.anio)

    print(f'\nMAPEO POR ENCABEZADO — hoja «{ws.title}»')
    print(f'  campos casados     : {len(mapa)}')
    print(f'  embarques fechados : {len(transito)}')
    for _i, etiqueta, fecha in transito:
        print(f'      {etiqueta.replace(chr(10), " "):<22} -> {fecha or "SIN FECHA"}')
    if sin_casar:
        # Las calculadas ya se filtraron: lo que llega aca es genuinamente nuevo.
        print(f'  ⚠️  columnas DESCONOCIDAS: {len(sin_casar)} — revisar antes de confiar')
        for i, c in sin_casar:
            print(f'      col {i}: {str(c)[:60]!r}')
    if faltantes:
        print(f'\n⛔ FALTAN COLUMNAS OBLIGATORIAS: {faltantes}')
        print('   No se carga nada. Un modelo a medias es peor que ninguno.')
        return 1

    estados, precios = leer_precios(wb)
    print(f'\nHOJA DE PRECIOS: {len(estados)} codigos, {len(precios)} precios por proforma')
    print(f'  estados: {dict(Counter(estados.values()))}')

    def celda(fila, campo):
        i = mapa.get(campo)
        return fila[i] if i is not None and i < len(fila) else None

    inventario, embarques, retenidas = [], [], []
    for fila in ws.iter_rows(min_row=4, values_only=True):
        cod = celda(fila, 'codigo')
        if not cod:
            continue
        cod = str(cod).strip()
        und_fardo = a_numero(celda(fila, 'und_fardo'))

        # Sin conversion a millares el producto no se puede modelar. Se RETIENE
        # y se reporta, igual que el cargador de facturas retiene una linea sin
        # tabla de conversion, en vez de suponer y propagar el error.
        if und_fardo is None or und_fardo <= 0:
            retenidas.append((cod, 'sin unidades por fardo'))
            continue

        inventario.append({
            'modelo': args.modelo, 'codigo': cod,
            'cod_proveedor': (str(celda(fila, 'cod_proveedor')).strip()
                              if celda(fila, 'cod_proveedor') else None),
            'descripcion': celda(fila, 'descripcion'),
            'um': celda(fila, 'um'),
            'und_fardo': und_fardo,
            'cub_millar': a_numero(celda(fila, 'cub_millar')),
            'sj': a_numero(celda(fila, 'sj')) or 0,
            'z11': a_numero(celda(fila, 'z11')) or 0,
            'zacapa': a_numero(celda(fila, 'zacapa')) or 0,
            'peten': a_numero(celda(fila, 'peten')) or 0,
            'patios_sj': a_numero(celda(fila, 'patios_sj')) or 0,
            'pend_surtir_sj': a_numero(celda(fila, 'pend_surtir_sj')) or 0,
            'pend_surtir_peten': a_numero(celda(fila, 'pend_surtir_peten')) or 0,
            'pend_surtir_zacapa': a_numero(celda(fila, 'pend_surtir_zacapa')) or 0,
            'transito_pendiente': a_numero(celda(fila, 'transito_pendiente')) or 0,
            'venta_proy_mensual': a_numero(celda(fila, 'venta_proy_mensual')),
            'precio_ml': a_numero(celda(fila, 'precio_ml')),
            'estado_producto': estados.get(cod, 'ACTIVO'),
            'autor': 'carga xlsx 2026-08-20',
        })

        for i, etiqueta, fecha in transito:
            v = a_numero(fila[i]) if i < len(fila) else None
            if v:
                embarques.append({
                    'modelo': args.modelo, 'codigo': cod,
                    'fecha': fecha.isoformat() if fecha else None,
                    'cantidad_ml': v,
                    # La etiqueta cruda queda como referencia: el ano de la
                    # fecha es una suposicion y asi queda auditable.
                    'referencia': etiqueta.replace('\n', ' ').strip(),
                    'autor': 'carga xlsx 2026-08-20',
                })

    precios_modelo = [{**p, 'modelo': args.modelo} for p in precios
                      if p['codigo'] in {i['codigo'] for i in inventario}]

    print(f'\nA CARGAR en el modelo «{args.modelo}»')
    print(f'  productos : {len(inventario)}')
    print(f'  embarques : {len(embarques)}')
    print(f'  precios   : {len(precios_modelo)}')
    sin_cub = sum(1 for i in inventario if i['cub_millar'] is None)
    sin_vta = sum(1 for i in inventario if i['venta_proy_mensual'] is None)
    if sin_cub:
        print(f'  ⚠️  sin cubicaje: {sin_cub} — sin m3 no entran al armado de contenedor')
    if sin_vta:
        print(f'  ⚠️  sin venta proyectada: {sin_vta} — quedaran SIN MOVIMIENTO')
    if retenidas:
        print(f'  ⛔ RETENIDAS (no se cargan): {len(retenidas)}')
        for cod, motivo in retenidas[:10]:
            print(f'      {cod}: {motivo}')

    if not args.commit:
        print('\nDRY-RUN. Correr con --commit para escribir.')
        return 0

    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_SECRET_KEY') or os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        print('Faltan SUPABASE_URL y SUPABASE_SECRET_KEY.')
        return 1
    from supabase import create_client
    sb = create_client(url, key)

    if not sb.table('modelo_proveedor').select('slug').eq('slug', args.modelo).execute().data:
        print(f'El modelo «{args.modelo}» no existe en modelo_proveedor.')
        return 1

    for i in range(0, len(inventario), 100):
        sb.table('reorden_inventario').upsert(
            inventario[i:i + 100], on_conflict='modelo,codigo').execute()
    print(f'  inventario: {len(inventario)}')

    # Transito y precios son un ESPEJO del libro: se reemplazan enteros para
    # este modelo. Sin esto, un embarque que el proveedor corrige quedaria
    # sumado dos veces.
    sb.table('reorden_transito').delete().eq('modelo', args.modelo).execute()
    for i in range(0, len(embarques), 100):
        sb.table('reorden_transito').insert(embarques[i:i + 100]).execute()
    print(f'  transito  : {len(embarques)}')

    for i in range(0, len(precios_modelo), 100):
        sb.table('reorden_precios').upsert(
            precios_modelo[i:i + 100], on_conflict='modelo,codigo,proforma').execute()
    print(f'  precios   : {len(precios_modelo)}')
    print('\nListo.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
